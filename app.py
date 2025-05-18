import os
import pandas as pd
import sqlite3
import streamlit as st
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import units

import zipfile

# CREAR TABLAS CON LAS COLUMNAS SELECCIONADAS
def guardar_tablas_en_bd(df1, df2, columnas_df1, columnas_df2):
    # Crear una conexión a la base de datos SQLite
    conn = sqlite3.connect('maestra.db')

    # Guardar las columnas del primer archivo en una tabla en la base de datos
    for columna in columnas_df1:
        if columna == 'nmalmcn':
            df1[columna] = df1[columna].astype(str)  # Convertir la columna en tipo de dato "texto"
        if columna == 'calmcn':
            df1[columna] = df1[columna].astype(str).str.zfill(5)
        if columna == 'fsrgstro':
            df1[columna] = pd.to_datetime(df1[columna]).dt.strftime('%Y-%m-%d')
        if columna == 'cartclo':
            df1[columna] = df1[columna].astype(str).str.zfill(14).str.replace(" ", "")
        if columna == 'nro_ruc_cli':
            df1[columna] = df1[columna].astype(str).str.replace('.0', '').str.replace(',', '')
    df1[columnas_df1].to_sql('movimientos', conn, if_exists='replace', index=False)


    for columna2 in columnas_df2:
        if columna2 == 'nmalmcn':
            df2[columna2] = df2[columna2].astype(str)  # Convertir la columna en tipo de dato "texto"
        if columna2 == 'calmcn':
            df2[columna2] = df2[columna2].astype(str).str.zfill(5)
        if columna2 == 'fvlote':
            df2[columna2] = pd.to_datetime(df2[columna2]).dt.strftime('%Y-%m-%d')
        if columna2 == 'cartclo':
            df2[columna2] = df2[columna2].str.slice(10).str.zfill(14).str.replace(" ", "")
        if columna2 == 'comprobante':
            df2[columna2] = df2[columna2].str.slice(5)

    # Guardar las columnas del segundo archivo en otra tabla en la base de datos
    df2[columnas_df2].to_sql('kardex', conn, if_exists='replace', index=False)



    # Cerrar la conexión a la base de datos
    conn.close()

# FIN CREAR TABLAS


# ELIMINAR TABLAS CON TODOS SUS DATOS
def borrar_datos_bd():
    # Conectar a la base de datos
    conn = sqlite3.connect('maestra.db')
    c = conn.cursor()

    # Obtener el nombre de todas las tablas en la base de datos
    c.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tablas = c.fetchall()
    tablas = [tabla[0] for tabla in tablas]

    # Borrar todas las tablas
    for tabla in tablas:
        c.execute(f"DROP TABLE {tabla}")

    # Cerrar la conexión a la base de datos
    conn.close()
# FIN ELIMINAR TABLAS


# CREA LA TABLA MAESTRA A TRAVES DE UNA CONSULTA
def realizar_consulta():
    # Conectar a la base de datos SQLite
    conn = sqlite3.connect('maestra.db')
    c = conn.cursor()

    # Ejecutar la consulta
    c.execute('''
     SELECT 'MAPFRE PERU' AS codigo, movimientos.calmcn, movimientos.cartclo, movimientos.dartclo, strftime('%d/%m/%Y', movimientos.fsrgstro), movimientos.gmvmnto, 
        movimientos.ndrcpcndo, movimientos.des_cli, movimientos.nro_ruc_cli, movimientos.cfmvmnto, movimientos.emfrccn, movimientos.tmalmcn, 
        movimientos.nmalmcn,movimientos.generico, movimientos.comercial, strftime('%d/%m/%Y',kardex.fvlote), kardex.lote, kardex.des_mov,'15-25° C' AS temperatura, julianday(kardex.fvlote) - julianday(movimientos.fsrgstro) AS  carta_canje, 
        CASE 
            WHEN julianday(kardex.fvlote) - julianday(movimientos.fsrgstro) < 366 THEN 'SI'
            ELSE 'NO'
        END AS carta,
        TRIM(
          REPLACE(
              movimientos.generico,
              movimientos.comercial,
              ''
          )
        ) AS presentacion
        FROM movimientos
        JOIN 
            kardex ON kardex.nmalmcn = movimientos.nmalmcn
            AND kardex.cartclo = movimientos.cartclo
            AND kardex.comprobante = movimientos.ndrcpcndo
            AND kardex.unid_in = movimientos.cfmvmnto
        WHERE movimientos.ndrcpcndo IS NOT NULL AND kardex.comprobante IS NOT NULL

    ''')

    # Obtener los resultados de la consulta
    resultados = c.fetchall()

    # Cerrar la conexión a la base de datos
    conn.close()

    return resultados
# FIN DE CREAR CONSULTA DE TABLA MAESTRA

# LIMPIA LAS FILAS PARA SEGUIR AÑADIENDO FACTURAS
def limpiar_hoja(hoja):
    hoja['E9'] = ''
    hoja['E10'] = ''
    hoja['C8'] = ''
    hoja['C6'] = ''
    for fila in range(15, 40):
        hoja[f'B{fila}'] = ''
        hoja[f'C{fila}'] = ''
        hoja[f'M{fila}'] = ''
        hoja[f'N{fila}'] = ''
        hoja[f'O{fila}'] = ''
        hoja[f'P{fila}'] = ''

def limpiar_hoja_guia(hoja):
    hoja['B2'] = ''
    hoja['B3'] = ''
    hoja['B4'] = ''
    hoja['B5'] = ''
    hoja['B6'] = ''
    hoja['B7'] = ''
    hoja['D3'] = ''
    hoja['D4'] = ''
    hoja['D5'] = ''
    hoja['D6'] = ''
    hoja['D7'] = ''

# FIN DE FILAS HOJAS

# CREAR LAS GUIAS CON LA INFORMACION DE LAS FACTURAS
def search_invoices(master_file, nro_facturas_lista):
    temp_dir = 'temp'
    os.makedirs(temp_dir, exist_ok=True)
    generated_files = []

    for nro_factura in nro_facturas_lista:
        df_master = pd.read_excel(master_file)
        filas_factura_con_valor_comun = df_master[df_master['ndrcpcndo'] == nro_factura]
        if filas_factura_con_valor_comun.empty:
            st.write(f"No se encontraron filas con el número de factura '{nro_factura}'")
        else:
            columnas_seleccionadas = filas_factura_con_valor_comun[
                ['ndrcpcndo', 'cartclo', 'dartclo', 'lote', 'fvlote', 'cfmvmnto',
                 'gmvmnto', 'des_cli', 'fsrgstro', 'temperatura']]
            wb = load_workbook('PE_HEA.xlsx')
            hoja = wb['impresion']

            hoja['E9'] = nro_factura
            hoja['E10'] = columnas_seleccionadas['gmvmnto'].iloc[0]
            hoja['C8'] = columnas_seleccionadas['des_cli'].iloc[0]
            hoja['C6'] = columnas_seleccionadas['fsrgstro'].iloc[0]

            fila = 15
            for index, row in columnas_seleccionadas.iterrows():
                hoja[f'B{fila}'] = str(row['cartclo']).zfill(14)
                hoja[f'C{fila}'] = row['dartclo']
                hoja[f'M{fila}'] = row['lote']
                hoja[f'N{fila}'] = row['fvlote']
                hoja[f'O{fila}'] = row['cfmvmnto']
                hoja[f'P{fila}'] = row['temperatura']
                fila += 1

            # Añadir la imagen de la firma
            firmas = [
                ('firma1.png', 'F48'),
                ('image4.png', 'B50'),
                ('image3.png', 'O48')  # Puedes añadir más si lo deseas
            ]

            # Añadir cada imagen en su posición
            for ruta_imagen, celda in firmas:
                firma_img = Image(ruta_imagen)
                firma_img.anchor = celda
                hoja.add_image(firma_img)

            # Guardar el archivo Excel
            nombre_archivo_excel = f'factura_{nro_factura}.xlsx'
            temp_file = os.path.join(temp_dir, nombre_archivo_excel)
            wb.save(temp_file)

            generated_files.append(temp_file)
            limpiar_hoja(hoja)

    if generated_files:
        zip_filename = 'facturas.zip'
        temp_zip_file = os.path.join(temp_dir, zip_filename)
        with zipfile.ZipFile(temp_zip_file, 'w') as zip_file:
            for file in generated_files:
                zip_file.write(file, os.path.basename(file))

        return temp_zip_file
    else:
        return ''



# CREAR LAS GUIAS CON LA INFORMACION DE LAS FACTURAS
def buscar_guias(master_file, nro_facturas_lista):
    temp_dir = 'temp'
    os.makedirs(temp_dir, exist_ok=True)
    generated_files = []
    df_master = pd.read_excel(master_file)

    for nro_factura in nro_facturas_lista:

        filas_factura_con_valor_comun = df_master[df_master['ndrcpcndo'] == nro_factura]


        if filas_factura_con_valor_comun.empty:
            st.write(f"No se encontraron filas con el número de factura '{nro_factura}'")
        else:

            for index, row in filas_factura_con_valor_comun.iterrows():
              columnas_seleccionadas = filas_factura_con_valor_comun[
                    ['ndrcpcndo', 'cartclo', 'lote', 'fvlote', 'cfmvmnto', 'des_cli', 'fsrgstro', 'comercial',
                     'generico',
                     'presentacion', 'nro_ruc_cli']]


              wb = load_workbook('LIBERACION_ACTA.xlsx')
              hoja = wb['impresion']

              hoja['B2'] = row['fsrgstro']
              hoja['B3'] = nro_factura
              hoja['B4'] = str(row['cartclo']).zfill(14).replace(" ", "")
              hoja['B5'] = row['lote']
              hoja['B6'] = row['fvlote']
              hoja['B7'] = row['presentacion']

              hoja['D3'] = row['comercial']
              hoja['D4'] = row['generico']
              hoja['D5'] = row['des_cli']
              hoja['D6'] = row['nro_ruc_cli']
              hoja['D7'] = row['cfmvmnto']

              template_wb = load_workbook('LIBERACION_ACTA.xlsx')
              template_hoja = template_wb['impresion']

              # Añadir la imagen de la firma
              firmas = [
                  ('firma1.png', 'F48'),
                  ('image4.png', 'B50'),
                  ('image3.png', 'O48')
                    # Puedes añadir más si lo deseas
              ]

              # Añadir cada imagen en su posición
              for ruta_imagen, celda in firmas:
                  firma_img = Image(ruta_imagen)
                  firma_img.anchor = celda
                  template_hoja.add_image(firma_img)

              nombre_archivo_excel = f'GUIA_LIBERACION_{nro_factura}_{index}.xlsx'
              temp_file = os.path.join(temp_dir, nombre_archivo_excel)
              wb.save(temp_file)

              generated_files.append(temp_file)
              limpiar_hoja_guia(hoja)

    if generated_files:
        zip_filename = 'facturas.zip'
        temp_zip_file = os.path.join(temp_dir, zip_filename)
        with zipfile.ZipFile(temp_zip_file, 'w') as zip_file:
            for file in generated_files:
                zip_file.write(file, os.path.basename(file))

        return temp_zip_file
    else:
        return ''





# FUNCION PARA MANEJAR LAS OPCIONES DE MENU
def handle_menu_option(option):
    if option == "CARGA A LA BASE DE DATOS (SUBIDA DE EXCELS)":
        st.title("Cargar archivos de Excel y guardar tablas en una base de datos")
        # Cargar los archivos de Excel mediante la interfaz de Streamlit
        archivo1 = st.file_uploader("Cargar archivo movimientos", type="xls")
        archivo2 = st.file_uploader("Cargar archivo kardex", type="xls")

        # Verificar si los archivos han sido cargados
        if archivo1 and archivo2:
            # Leer los archivos de Excel en DataFrames
            df1 = pd.read_excel(archivo1)
            df2 = pd.read_excel(archivo2)

            # Seleccionar las columnas de interés mediante la interfaz de Streamlit
            columnas_df1 = st.multiselect("Seleccionar columnas del archivo 1", df1.columns)
            columnas_df2 = st.multiselect("Seleccionar columnas del archivo 2", df2.columns)

            # Verificar si se han seleccionado columnas en ambos archivos
            if columnas_df1 and columnas_df2:
                # Mostrar un botón para guardar las tablas en la base de datos
                if st.button("Guardar tablas en la base de datos"):
                    guardar_tablas_en_bd(df1, df2, columnas_df1, columnas_df2)
                    st.success("Las tablas se han guardado exitosamente en la base de datos.")
            else:
                st.warning("Debes seleccionar al menos una columna en ambos archivos.")

        # Código para la opción 1
        #st.write("Estás en la Opción 1")
    elif option == "VISTA TABLAS MOVIMIENTOS Y KARDEX":
        #st.title("Opción 2")

        st.title("Consulta TABLA")

        # Ruta de la base de datos
        db_path = os.path.join('maestra.db')

        # Conexión a la base de datos
        conn = sqlite3.connect(db_path)
        c = conn.cursor()

        # Obtener el nombre de todas las tablas en la base de datos
        c.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tablas = c.fetchall()
        tablas = [tabla[0] for tabla in tablas]

        # Widget para seleccionar la tabla
        tabla_seleccionada = st.selectbox("Seleccionar tabla", tablas)

        if tabla_seleccionada:
            # Consultar todos los datos de la tabla seleccionada
            consulta = f"SELECT * FROM {tabla_seleccionada}"
            c.execute(consulta)
            resultados = c.fetchall()

            if resultados:
                # Crear un DataFrame con los resultados de la consulta
                columnas_df = [descripcion[0] for descripcion in c.description]
                datos_df = pd.DataFrame(resultados, columns=columnas_df)

                # Mostrar los datos en Streamlit
                st.write(datos_df)
            else:
                st.write("No se encontraron datos en la tabla seleccionada.")

        conn.close()

        # Mostrar un botón para borrar todos los datos de la base de datos
        if st.button("Borrar todos los datos de la base de datos"):
            borrar_datos_bd()
            st.success("Se han borrado todos los datos, columnas y la tabla de la base de datos.")

        # Código para la opción 2
        #st.write("Estás en la Opción 2")
    elif option == "OBTENER MAESTRA.CSV -> EXCEL":
        # st.title("Opción 3")
        # Código para la opción 3
        # Configurar la página Streamlit
        st.title("Consulta de coincidencias")

        # Botón para realizar la consulta
        if st.button("Consultar las coincidencias"):
            # Realizar la consulta
            resultados = realizar_consulta()

            # Mostrar los resultados en una tabla
            if len(resultados) > 0:
                st.write("Resultados:")
                df = pd.DataFrame(resultados,
                                  columns=["codigo", "calmcn", "cartclo", "dartclo", "fsrgstro", "gmvmnto", "ndrcpcndo",
                                           "des_cli", "nro_ruc_cli", "cfmvmnto", "emfrccn", "tmalmcn", "nmalmcn", "generico",
                                           "comercial", "fvlote", "lote", "des_mov", "temperatura", "carta_canje",
                                           "carta","presentacion"])
                st.dataframe(df)




            else:
                st.write("No se encontraron resultados.")
        #st.write("Estás en la Opción 3")

    elif option == "CREAR GUIA DE FACTURAS":

        st.title('CREACION DE GUIA DE FACTURAS')

        uploaded_master_file = st.file_uploader("Upload Master File", type=["xlsx"])
        nro_facturas = st.text_area("Enter Invoice Numbers (comma-separated)")

        if st.button('Search and Export'):
            if uploaded_master_file is not None and nro_facturas:
                nro_facturas_lista = [f.strip() for f in nro_facturas.split(',')]
                temp_zip_file = search_invoices(uploaded_master_file, nro_facturas_lista)

                if temp_zip_file:
                    with open(temp_zip_file, 'rb') as f:
                        st.download_button(
                            label="Download ZIP",
                            data=f,
                            file_name='facturas.zip',
                            mime='application/zip'
                        )
                else:
                    st.write("No invoices found or generated.")
            else:
                st.write("Please upload a master file and enter invoice numbers.")


    elif option == "ACTA DE LIBERACION":

        st.title('CREACION DE ACTAS DE LIBERACION')

        uploaded_master_file = st.file_uploader("Upload Master File", type=["xlsx"])
        nro_facturas = st.text_area("Enter Invoice Numbers (comma-separated)")

        if st.button('BUSCAR Y EXPORTAR'):
            if uploaded_master_file is not None and nro_facturas:
                nro_facturas_lista = [f.strip() for f in nro_facturas.split(',')]
                temp_zip_file = buscar_guias(uploaded_master_file, nro_facturas_lista)

                if temp_zip_file:
                    with open(temp_zip_file, 'rb') as f:
                        st.download_button(
                            label="Download ZIP",
                            data=f,
                            file_name='guia_liberacion.zip',
                            mime='application/zip'
                        )
                else:
                    st.write("No invoices found or generated.")
            else:
                st.write("Please upload a master file and enter invoice numbers.")


# Crear el menú en la barra lateral
st.sidebar.title("Menú")
menu_options = ["CARGA A LA BASE DE DATOS (SUBIDA DE EXCELS)", "VISTA TABLAS MOVIMIENTOS Y KARDEX", "OBTENER MAESTRA.CSV -> EXCEL", "CREAR GUIA DE FACTURAS","ACTA DE LIBERACION"]
selected_option = st.sidebar.radio("Selecciona una opción:", menu_options, index=0)

# Llamar a la función correspondiente al menú seleccionado
handle_menu_option(selected_option)
